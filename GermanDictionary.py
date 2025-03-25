import random
import openpyxl

correct = 0
dict_file = openpyxl.load_workbook(
    "Copia de Hoja de cálculo sin título - 14 de agosto, 15_40.xlsx"
)
word_list = dict_file["Hoja 1"]


def main():
    answer = input(
        "HELLO! this is 1000 Words in German! \n"
        "In this game, you will receive a random german word, and you have to write the translation! \n"
        "If you run out of lives, you loose! so be careful! \n"
        "Enter /y to start from the beggining, or enter 's' to pick a level!\n "
    )

    words1 = {}
    words2 = {}
    words3 = {}
    words4 = {}
    words5 = {}
    words6 = {}
    words7 = {}
    words8 = {}
    words9 = {}
    words10 = {}

    def dict_making(start, finish, empty_dict):
        for word_row in range(start, finish):
            german_word = word_list.cell(word_row, 1).value
            words_values = word_list.cell(word_row, 4).value
            words_values2 = word_list.cell(word_row, 5).value
            words_values3 = word_list.cell(word_row, 6).value
            empty_dict[german_word] = [words_values, words_values2, words_values3]

    dict_making(1, 101, words1)
    dict_making(101, 201, words2)
    dict_making(201, 301, words3)
    dict_making(301, 401, words4)
    dict_making(401, 501, words5)
    dict_making(501, 601, words6)
    dict_making(601, 701, words7)
    dict_making(701, 801, words8)
    dict_making(801, 901, words9)
    dict_making(901, 1001, words10)

    level1 = list(words1.keys())
    level2 = list(words2.keys())
    level3 = list(words3.keys())
    level4 = list(words4.keys())
    level5 = list(words5.keys())
    level6 = list(words6.keys())
    level7 = list(words7.keys())
    level8 = list(words8.keys())
    level9 = list(words9.keys())
    level10 = list(words9.keys())

    random.shuffle(level1)
    random.shuffle(level2)
    random.shuffle(level3)
    random.shuffle(level4)
    random.shuffle(level5)
    random.shuffle(level6)
    random.shuffle(level7)
    random.shuffle(level8)
    random.shuffle(level9)
    random.shuffle(level10)

    def play_game(word_keys, word_dict, level):
        global correct
        if level == 1:
            print("you are in level: 1")
        if level >= 2:
            print("you are in level " + str(level) + ". Keep going!")
        for keyword in word_keys:
            attempts = 5
            print(" ")
            print(keyword)
            print(" ")
            userInputAnswer = input("ANSWER: ")
            if userInputAnswer == "s":
                print(word_dict[keyword])

            if userInputAnswer in (word_dict[keyword]):
                print("That is correct!")
                print(" ")
                correct += 1
                print("correct: " + str(correct))
                print(" ")
                print("-" * 25)

            while userInputAnswer not in (word_dict[keyword]):
                if userInputAnswer == "s":
                    print(word_dict[keyword])

                print("not quite!")
                print(" ")
                attempts -= 1
                print("correct: " + str(correct))
                print("Attepmts: " + str(attempts))
                print(" ")

                print(keyword)
                print(" ")
                userInputAnswer = input("ANSWER: ")

                if userInputAnswer in (word_dict[keyword]):
                    print("That is correct!")
                    print(" ")
                    correct += 1
                    print("correct: " + str(correct))
                    print(" ")
                    print("-" * 25)

                if attempts == 0:
                    print(keyword)
                    restart = input(
                        "You lost! do you want to try again? y/ to continue, n/ to exit!"
                    )
                    if restart == "y":
                        correct = 0
                        main()
                    else:
                        exit()

    while answer == "s":
        user_input = input("Select a level from 1 to 10 \n")
        while user_input == "1":
            play_game(level1, words1, 1)
            play_game(level2, words2, 2)
            play_game(level3, words3, 3)
            play_game(level4, words4, 4)
            play_game(level5, words5, 5)
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "2":
            play_game(level2, words2, 2)
            play_game(level3, words3, 3)
            play_game(level4, words4, 4)
            play_game(level5, words5, 5)
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "3":
            play_game(level3, words3, 3)
            play_game(level4, words4, 4)
            play_game(level5, words5, 5)
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "4":
            play_game(level4, words4, 4)
            play_game(level5, words5, 5)
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "5":
            play_game(level5, words5, 5)
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "6":
            play_game(level6, words6, 6)
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "7":
            play_game(level7, words7, 7)
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "8":
            play_game(level8, words8, 8)
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "9":
            play_game(level9, words9, 9)
            play_game(level10, words10, 10)
        while user_input == "10":
            play_game(level10, words10, 10)

    while answer == "y":
        play_game(level1, words1, 1)
        play_game(level2, words2, 2)
        play_game(level3, words3, 3)
        play_game(level4, words4, 4)
        play_game(level5, words5, 5)
        play_game(level6, words6, 6)
        play_game(level7, words7, 7)
        play_game(level8, words8, 8)
        play_game(level9, words9, 9)
        play_game(level10, words10, 10)


main()
